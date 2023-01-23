from multiprocessing import Semaphore
import numpy as np
import pandas as pd
import re
from sklearn.model_selection import train_test_split
import openpyxl
from openpyxl.styles import PatternFill, Font
import itertools
import xlsxwriter
import PIL.Image
import math
import os
from colorama import Fore, Style
import matplotlib.pyplot as plt
import datetime

param_types = {pd.DataFrame: ['X_train', 'X_oos','X_oot'],
                pd.Series: ['y_train', 'y_oos','y_oot'], 
                str: ['pipeline'] }


def printcolor(text, color = 'k', size=13, weight = None):

    #plt.figure(figsize =(0.05, 0.05))
    #plt.axis('off')
    #plt.title(text, color = color, size = size, weight=weight)
    #plt.show()
    colors = {'r':'\033[31m', 'g':'\033[32m','y':'\033[33m', 'b':'\033[34m', 'k':''}
    style = {'bold':'\033[1m', None:'' }
    print(colors[color]+style[weight]+"{}\033[0m".format(text))


def table_style(color_list):
    txt_clr = lambda val: "#9c0031" if val == "Red" else "#9c5700" if val == "Yellow" else "#006100" if val == "Green" else None
    bg_clr = lambda val: "#ffc7ce" if val == "Red" else "#ffeb9c" if val == "Yellow" else "#c6efce" if val == "Green" else None
    style = [{"index": num,
              "color": txt_clr(val),
              "backgroundColor": bg_clr(val)
              } for num, val in enumerate(color_list)]
    return style


def make_test_report(color=None, table=None, title='', color_list = [None], name=__file__,short_res='',long_res='' ):
    result_test = {
        "result": 
            {"semaphore": 
                {"color": color,
                 "title": title},
                 "table": table,
                 "tablestyle": table_style(color_list)},
            "result_out": pd.DataFrame({
                'name': ['metric_stability_test '], 'color':[color],
                'short_result':[short_res[:-1]], 'long_result':[long_res],
                'time':[str(datetime.datetime.today())]})}
    return result_test



def mincolor(colors):
    #возвращает худший цвет сетофора
    allc = ["Red", "Yellow", "Green"]
    for c in allc:
        if c in set(colors):
            return c
    # Если у нас нет блока/нет неинформативных тестов, то возвращаем None по нему
    return None

def semafore_agregation(table_test, type_agg = 'standart'):
    semaphore = None
    if type_agg=='standart':
        return mincolor(list(table_test['results'].apply(lambda x: x['result']['semaphore']['color']).values)) #min по всем тестам

    elif  type_agg=='modify_stability_metric':
        stab_metric = table_test[table_test['name_test']=='metric_stability_test']['results'].values[0]['result']['semaphore']['color']
        #del dict_color_test['metric_stability_test']
        other_metrics = mincolor(list(table_test[table_test['name_test']!='metric_stability_test']['results'].apply(lambda x: x['result']['semaphore']['color']).values))
        if other_metrics in ['Green', 'Yellow']:
            return stab_metric
        else:
            if  stab_metric=='Green':
                return 'Yellow'
            else:
                return 'Red'

    elif type_agg=='modify_quality_data':
        block_res = table_test.groupby(by = 'block')['results'].agg(lambda x: mincolor(list(map(lambda y: y['result']['semaphore']['color'],x))))
        block_res.reset_index(inplace = True)
        block_res.columns = ['block', 'color']
        quality_metrics = list(block_res[block_res['block']=='quality_data']['color'].values)[0]
        other_metrics = mincolor(list(block_res[block_res['block']!='quality_data']['color'].values))
        if other_metrics in ['Green', 'Yellow']:
            return quality_metrics
        else:
            if  quality_metrics=='Green':
                return 'Yellow'
            else:
                return 'Red'
    else:
        pass
    return semaphore
    


def sortstrcallabele(somelist):
    # осртирует объекты разных типов. Создано для сортировки списка из str и callabel элементов.
    # str объекты сортируются сначала, после них сортируются  callable
    listO_of_str = list(filter(None, list(map(lambda x: x if type(x)==str else None, somelist))))
    listO_of_no_str = list(filter(None, list(map(lambda x: x if type(x)!=str else None, somelist))))
    listO_of_str.sort()
    listO_of_no_str.sort(key= lambda x: x.__name__)
    return listO_of_str+listO_of_no_str
    
    
def agg_two_criterion(abs_light, rel_light):
    if (abs_light == 'Red') and (rel_light == 'Red'):
        test_resultl = 'Red'
    elif (abs_light == 'Green') or (rel_light == 'Green'):
        test_result = 'Green'
    else:
        test_result ='Yellow'
    return test_result

def numpy_nan_to_none(val):
    try:
        if np.isnan(val):
            return None
    except TypeError:
        pass
    return val
    
def parse_params(args):
    if numpy_nan_to_none(args) is None:
        return {}
    params = [i.strip().split('=') for i in args.split('\n')]
    params = {i[0].strip(): eval(i[1].strip()) for i in params}
    return params
    
def validation_result_reversed(value, boundaries):
    """
    Determine colour of test from value and boundaries. Lesser value means better result.

    :param value: test value. Format: float
    :param boundaries: thresholds for test. Format: (yellow_value, red_value)
    :return: string colour: 'Green', 'Yellow' or 'Red'.
    """
    if value <= min(boundaries):
        return 'Green'
    elif value > max(boundaries):
        return 'Red'
    else:
        return 'Yellow'
        
        
def validation_result(value, boundaries):
    """
    Determine colour of test from value and boundaries. Higher value means better result.

    :param value: test value. Format: float
    :param boundaries: thresholds for test. Format: (yellow_value, red_value)
    :return: string colour: 'Green', 'Yellow' or 'Red'.
    """
    if value <= min(boundaries):
        return 'Red'
    elif value > max(boundaries):
        return 'Green'
    else:
        return 'Yellow'
        
        
def clear_string(string):
    """
    Clear string from symbols restricted to use in Windows filename

    :param string: input string
    :return: cleared string
    """
    string = str(string)
    return re.sub('[/\\\\:*?"<>|.]', ' ', string)


def mean_color(color_list):
    pass
    
    
 
    
def train_test_split_upgraded(*args, split_by_index=False, split_by_columns=False, **kwargs):
    if (not split_by_index) and (not split_by_columns):
        return train_test_split(*args, **kwargs)
    elif split_by_index:
        indices = train_test_split(*[sorted(set(i.index)) for i in args], **kwargs)
        return list(itertools.chain.from_iterable((arr.loc[indices[2*num]],
                                                   arr.loc[indices[2*num+1]]) for num, arr in enumerate(args)))
    elif split_by_columns:
        val_cols = train_test_split(*[sorted(set(i[split_by_columns].values)) for i in args], **kwargs)
        return list(itertools.chain.from_iterable((arr.loc[val_cols[2*num]],
                                                   arr.loc[val_cols[2*num+1]]) for num, arr in enumerate(args)))
    return None

def pass_none_values(func):
    """Wrap the function to do nothing if the first parameter is None of any kind"""
    def function_wrapper(*args, **kwargs):
        # Since we expect text input at second position (the first is self), we hard code args[1] check

        # Screw numpy NaN...
        if numpy_nan_to_none(args[1]) is None:
            return -1

        return func(*args, **kwargs)
    return function_wrapper

class XlsxBot:
    DEFAULT_CELL_WIDTH = 8.43  # 64 pixels - default Excel cell width
    DEFAULT_CELL_PIXELS = 64
    PIXEL_TO_WIDTH_RATIO = DEFAULT_CELL_WIDTH / DEFAULT_CELL_PIXELS

    ROW_HEIGHT = 20
    CHAR_WIDTH_PIXELS = 8 * 1.125
    GLOBAL_WIDTH_PIXELS = 800
    MAX_CELL_WIDTH_PIXELS = 150 * 2

    MAX_CELL_WIDTH = MAX_CELL_WIDTH_PIXELS / DEFAULT_CELL_WIDTH
    MIN_CELL_WIDTH = DEFAULT_CELL_WIDTH
    # CHAR_HEIGHT = 20

    USE_MILOVSKY_COEF = False  # Set to True so that inserted pictures match excel pixel size (in some cases)
    MILOVSKY_COEF_SCALING = 1.38  # Excel pixels to PNG pixels

    def __init__(self, excel_path, sheet_name='Sheet1', **kwargs):
        self.excel_path = excel_path
        self.wb = xlsxwriter.Workbook(excel_path)
        self.wb.nan_inf_to_errors = True
        self.formats = {}
        self.sheet = None
        self.row = 0
        self.col = 0

        self.GLOBAL_WIDTH_PIXELS = kwargs.get("global_width", 800)
        self.MAX_CELL_WIDTH_PIXELS = kwargs.get("max_col_width", 150)
        self.MAX_CELL_WIDTH = self.MAX_CELL_WIDTH_PIXELS / self.DEFAULT_CELL_WIDTH
        self.USE_MILOVSKY_COEF = kwargs.get("milovsky_coef", False)

        self.set_formats(self.generate_formats())
        self.set_active_sheet(sheet_name)

    def generate_formats(self):
        num_formats = {
            'pcnt': {'num_format': '0.00%'},
            'float': {'num_format': '0.000'},
            'int': {'num_format': '0'},
            None: {}
        }
        colors = {
            'green': {'bg_color': '#C6EFCE', 'font_color': '#006100'},
            'yellow': {'bg_color': '#FFEB9C', 'font_color': '#9C5700'},
            'red': {'bg_color': '#FFC7CE', 'font_color': '#9C0006'},
            None: {}
        }
        borders = ['bottom', 'right', 'bottom_right', None]

        p = itertools.product(num_formats, colors, borders)
        formats = {(num, clr, bord): self.wb.add_format({**num_formats[num], **colors[clr]}) for num, clr, bord in p}

        # Bad implementation, but works for now
        for f in formats:
            if 'bottom' in f:
                formats[f].set_bottom()
            if 'right' in f:
                formats[f].set_right()
            if 'bottom_right' in f:
                formats[f].set_bottom()
                formats[f].set_right()

        formats['table_title'] = self.wb.add_format({'bold': True, 'border': True, 'align': 'center',
                                                     'valign': 'vcenter', 'text_wrap': 1})
        # formats['table_index'] = self.wb.add_format({'bold': True, 'border': True})
        formats['sheet_header'] = self.wb.add_format({'bold': True, 'font_size': 14})
        formats['local_header'] = self.wb.add_format({'bold': True, 'italic': True})
        formats['textbox'] = self.wb.add_format({})

        formats['red_threshold'] = self.wb.add_format({'bold': True, 'font_color': '#C00000', 'bg_color': 'white'})
        formats['yellow_threshold'] = self.wb.add_format({'bold': True, 'font_color': '#FFC000', 'bg_color': 'white'})
        formats['green_threshold'] = self.wb.add_format({'bold': True, 'font_color': '#006100', 'bg_color': 'white'})

        formats['cell_default'] = self.wb.add_format({'bg_color': 'white'})  # TODO: merge with other formats
        return formats

    def get_col_formats(self, df):
        col_formats = []
        for c in df.keys():
            fmt = 'pcnt' if '%' in c \
                else 'float' if df[c].dtype == np.float64 \
                else 'int' if df[c].dtype == np.int64 \
                else None
            col_formats.append(fmt)
        return col_formats

    def set_active_sheet(self, name):
        if name not in self.wb.sheetnames:
            self.sheet = self.wb.add_worksheet(name)

            # Fill *kind of* all cells with default cell format
            for c in range(26):
                # Manually set one-letter columns to prevent overwriting the whole range later:
                self.sheet.set_column(c, c, self.DEFAULT_CELL_WIDTH, self.formats.get('cell_default'))
            self.sheet.set_column('AA:ZZ', self.DEFAULT_CELL_WIDTH, self.formats.get('cell_default'))

        self.sheet = self.wb.sheetnames[name]
        self.row = 0
        self.col = 0

    def set_formats(self, f):
        self.formats = f

    def set_cursor(self, row=0, col=0):
        self.row = row
        self.col = col

    def calc_text_len(self, text):
        text_len = len(text) * self.CHAR_WIDTH_PIXELS * self.PIXEL_TO_WIDTH_RATIO
        return min(self.MAX_CELL_WIDTH, text_len)

    def adjust_col_width(self, col_ind, text):
        text_len = self.calc_text_len(text)
        width = max(self.MIN_CELL_WIDTH, text_len)
        self.sheet.set_column(col_ind, col_ind, width, self.formats.get('cell_default'))

    def insert_table(self, df):
        col_formats = self.get_col_formats(df)

        # If one attempts to insert second table to list, index width setup will overwrite the first one
        index_len = max(self.calc_text_len(str(value)) for value in df.index) * 2
        self.sheet.set_column(self.col, self.col, index_len, self.formats.get('cell_default'))

        for i, colname in enumerate(df.columns):
            if self.sheet.col_sizes.get(i+1, self.DEFAULT_CELL_WIDTH) == self.DEFAULT_CELL_WIDTH:
                self.adjust_col_width(i+1, colname)
            self.sheet.write(self.row, self.col + i + 1, colname, self.formats.get('table_title'))

        for i, row_content in enumerate(df.iterrows()):
            index, values = row_content
            row_color = values.get("Результат теста")
            row_color = row_color.lower() if row_color else row_color

            self.sheet.write(self.row + i + 1, self.col, index, self.formats.get('table_title'))
            for j, val in enumerate(values):
                col_format = col_formats[j]

                bord_format = []
                if i == len(df) - 1:
                    bord_format.append('bottom')
                if j == len(values) - 1:
                    bord_format.append('right')
                bord_format = '_'.join(bord_format) if len(bord_format) > 0 else None

                self.sheet.write(self.row + i + 1, self.col + j + 1, numpy_nan_to_none(val),
                                 self.formats.get((col_format, row_color, bord_format)))

        self.row += len(df) + 2

    def insert_image(self, img_path, shift=None):
        x, y = PIL.Image.open(img_path).size

        # Convert x and y to Excel pixel size
        if self.USE_MILOVSKY_COEF:
            x *= self.MILOVSKY_COEF_SCALING
            y *= self.MILOVSKY_COEF_SCALING

        scaling = 1
        if x > self.GLOBAL_WIDTH_PIXELS:
            scaling = self.GLOBAL_WIDTH_PIXELS / x

        self.sheet.insert_image(self.row, self.col + 1, img_path, {'x_scale': scaling, 'y_scale': scaling})
        if shift is None:
            shift = math.ceil(y * scaling / self.ROW_HEIGHT) + 1
        self.row += shift

    @pass_none_values
    def insert_text(self, string, fmt=None, shift=2):
        self.sheet.write(self.row, self.col + 1, string, fmt)
        self.row += shift

    @pass_none_values
    def insert_textbox(self, text):
        text_height = math.ceil(len(text) * self.CHAR_WIDTH_PIXELS / self.GLOBAL_WIDTH_PIXELS) + 1
        self.sheet.insert_textbox(self.row, self.col + 1, text, {'width': self.GLOBAL_WIDTH_PIXELS,
                                                                 'height': text_height * self.ROW_HEIGHT})
        self.row += text_height + 1

    @pass_none_values
    def insert_titled_text(self, text, title):
        self.insert_text(title, self.formats.get('local_header'), shift=1)
        self.insert_textbox(text)

    def insert_title(self, title):
        self.insert_text(title, self.formats.get('sheet_header'))

    def insert_thresholds_info(self, red, yellow, green):
        if numpy_nan_to_none(red) is None and numpy_nan_to_none(yellow) is None and numpy_nan_to_none(green) is None:
            return -1
        self.insert_text("Границы светофора:", self.formats.get('local_header'), shift=1)
        self.insert_text(red, self.formats.get('red_threshold'), shift=1)
        self.insert_text(yellow, self.formats.get('yellow_threshold'), shift=1)
        self.insert_text(green, self.formats.get('green_threshold'), shift=2)

    def save(self):
        self.wb.close()
        """
        Closes workbook handler to save inserted data, then reopens it.
        Warning! It will OVERWRITE the file if saved again.
        # self.__init__(self.excel_path, self.sheet.name)
        """


def agg_results(name='Results.xlsx',
                global_width=800,
                max_col_width=150,
                milovsky_coef=False,
                config=None,
                decrease_model_risk=False,
                path='.',
                no_results_folder=False):
    """
    Aggregate resulting xlsx files into one.

    :param name: name of output aggregated file
    :param global_width: maximum plot / textbox width
    :param max_col_width: maximum column width
    :param milovsky_coef: turn on plot scaling (makes plots smaller, sometimes helps in case of unsuitable plots)
    :param config: test descriptions and thresholds
    :param decrease_model_risk: do not include PDP plots in aggregated file, decreasing its size (and model risk as well)
    :param path: relative or absolute path for dump data
    :param no_results_folder: do not append 'results' to path
    :return: 1
    """

    # Close Results.xlsx if opened - temporarily disabled
    # try: win32.gencache.EnsureDispatch('Excel.Application').Workbooks(name).Close()
    # except: pass

    excel = None
    if config is None:
        config = {}

    # Get file name
    results_path = path
    agg_file_path = os.path.join(results_path, name)

    # Iterate through all the xlsx test outputs
    for file in sorted(os.listdir(results_path)):
        if '.xlsx' in file and file != name:
            # Get the numeric name of test
            test_name = file[:-5]

            # Init Excel writer if it hasn't been initialized yet
            if excel is None:
                excel = XlsxBot(agg_file_path, test_name, global_width=global_width, max_col_width=max_col_width,
                                milovsky_coef=milovsky_coef)
            else:
                excel.set_active_sheet(test_name)

            # Get the test info
            test_info = config.get(test_name, {})

            # Textbox insertion causes a bug that totally breaks all the images in file.
            # Fix that prevents it: we insert a blank image into each excel sheet.
            excel.insert_image(os.path.join(os.path.dirname(__file__), "blank.png"), shift=0)

            # Print test title
            test_title = test_info.get('Название', np.nan)
            excel.insert_title(test_title)

            # Print test result to file
            data = pd.read_excel(os.path.join(results_path, file), index_col=0)
            excel.insert_table(data)

            # Print test thresholds
            red_threshold = test_info.get('Границы красный', np.nan)
            yellow_threshold = test_info.get('Границы желтый', np.nan)
            green_threshold = test_info.get('Границы зеленый', np.nan)
            excel.insert_thresholds_info(red_threshold, yellow_threshold, green_threshold)

            # Not implemented yet #
            """
                # Format error traceback
                if column_name == 'Traceback':
                    worksheet[excel_column + '2'].alignment = openpyxl.styles.Alignment(wrap_text=True)
            """
            # Insert pictures into excel
            for pic_name in os.listdir(path):
                if '.png' in pic_name and test_name in pic_name:
                    # CRUTCH
                    if decrease_model_risk and 'PDP' in pic_name: continue
                    excel.insert_image(os.path.join(path, pic_name))

            # Print test purpose if it is defined
            test_purpose = test_info.get('Цель', np.nan)
            excel.insert_titled_text(test_purpose, 'Цель теста:')

            # Print test interpretation if it is defined
            test_interpretation = test_info.get('Интерпретация', np.nan)
            excel.insert_titled_text(test_interpretation, 'Интерпретация:')

            # Not implemented yet #
            """
            # Format execution time list
            if sheet_name == "Execution_time":
                for i in [str(i) for i in range(1, len(data)+2)]:
                    for col in "BCDE":
                        if worksheet["E" + i].value == "OK":
                            worksheet[col + i].fill = PatternFill(start_color="00C6EFCE", fill_type='solid')
                            worksheet[col + i].font = Font(color="00006100")
                        elif worksheet["E" + i].value == "Error":
                            worksheet[col + i].fill = PatternFill(start_color="00FFC7CE", fill_type='solid')
                            worksheet[col + i].font = Font(color="009C0006")
            """
    excel.save()
    del excel

    return 1



